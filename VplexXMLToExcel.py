"""
 VplexXMLToExcel

 Transforme le XML de configuration d'un VPLEX en fichier XLS

 Author : Olivier Guyot

 Developped during off hours (vacations)
 Code is licences under GNU GPL v3

 TODO :
    TOUT

"""
import logging.config
import argparse
import xml.etree.ElementTree as ET
from shutil import copyfile
import time
import openpyxl
import math


"""
Initialize logging


File SymApiToExcel.logging mus be present and describe all availables logging configuration
Per default in the file :
2 loggers : 1 for console; one in file. (no rotation).
File is more for debug purposes.
In case of issue, truncate the log file or upgrade the logger level to Fatal

"""
logging.config.fileConfig('SymApiToExcel.logging')
logger = logging.getLogger('root')


class mesObjets:
    """
    meObjets : mother class for all custom objects.

    you will find static methoods for code refactoring (why do it 10 times when you can write once and call many (find / findall)

    You will also find toString methods that transform an object to string listing all attributes to string.
    if an attribute is a list, then data are not printed.
    """


    def toString(self) -> str:
        """
        toString : put in a nice string all attributes of an object.

        :return: A string cotaining all attributes and values of an object
        """
        result = ""
        variables = self.__dict__.items()
        for variable, value in variables:
            if (str(variable).startswith("list_")):
                result = result + variable + " ====> LIST " + str(len(list(value))) + " Elements \n"
            else:
                result = result + variable + "  ====>  " + str(value) + "\n"
        return result

    def getValue(self, findV):
        """
        retrieve the value of a variable inside an object
        :param findV:
        :return: the value
        """
        variables = self.__dict__.items()
        for variable, value in variables:
            if variable == findV:
                return value
        return

class vplexPort(mesObjets):
    PortName = ""
    PortWWN = ""
    NumberExportedVolumes = 0


    @staticmethod
    def loadFromXML(XMLString):
        port = vplexPort()
        port.PortName = XMLString.find("PortName").text.strip()
        port.PortWWN = XMLString.find("PortWWN").text.strip()
        port.NumberExportedVolumes = int(XMLString.find("NumberExportedVolumes").text)

        return port


class storageArray(mesObjets):
    VendorID = ""
    ProductID = ""
    Revision = ""
    ArrayID = ""
    FailoverMode = ""
    NumberPaths = 0
    nbVolume = 0
    nbVolumeClaim = 0


    @staticmethod
    def loadFromXML(XMLString):
        baie = storageArray()
        baie.VendorID = XMLString.find("VendorID").text.strip()
        baie.ProductID = XMLString.find("ProductID").text.strip()
        baie.Revision = XMLString.find("Revision").text.strip()
        baie.ArrayID = XMLString.find("ArrayID").text.strip()
        baie.FailoverMode = XMLString.find("FailoverMode").text.strip()
        baie.NumberPaths = int(XMLString.find("NumberPaths").text)
        todo = XMLString.find("StorageElements")
        baie.nbVolume = int(todo.find("NumberSEs").text)
        baie.nbVolumeClaim = int(todo.find("NumberClaimedSEs").text)

        return baie


class vplex(mesObjets):
    clusterTLA = ""
    clusterId = ""
    clusterNumber = 0
    directorCount = 0
    engineCount = 0
    operationalStatus = ""
    healthState = ""
    healthIndications = ""
    list_storageA = []
    list_feports = []
    list_feportsA = []
    list_feportsB = []
    NumberViews = 0
    NumberInitiatorPorts = 0
    ClaimedCapacity = ""
    StorageVolumes = 0
    nbRaid0Devices = 0
    nbRaidCDevices = 0
    nbRaid1Devices = 0
    nbDistributedDevices = 0
    nbRemoteFrom = 0
    nbExportedVV = 0
    nbLocalCG = 0
    nbDistributedCG = 0
    modelType = ""
    seed_id = ""
    version = ""
    siteID = ""




    @staticmethod
    def loadFromXML(XMLString):
        MyVPlex = vplex()
        ClusterAttributes = XMLString.find("ClusterAttributes")
        MyVPlex.clusterTLA = ClusterAttributes.find("clusterTLA").text.strip()
        MyVPlex.clusterId = ClusterAttributes.find("cluster-id").text.strip()
        MyVPlex.clusterNumber = int(ClusterAttributes.find("cluster-number").text)
        MyVPlex.directorCount = int(ClusterAttributes.find("director-count").text)
        MyVPlex.engineCount = int(MyVPlex.directorCount/2)
        MyVPlex.operationalStatus = ClusterAttributes.find("operational-status").text.strip()
        MyVPlex.healthState=ClusterAttributes.find("health-state").text.strip()
        MyVPlex.healthIndications = ClusterAttributes.find("health-indications").text.strip()


        views = XMLString.find("Views")

        MyVPlex.NumberViews = views.find("NumberViews").text.strip()
        MyVPlex.NumberInitiatorPorts = views.find("NumberInitiatorPorts").text.strip()

        portlist = views.find("PortList")
        MyVPlex.list_feports = []
        MyVPlex.list_feportsA = []
        MyVPlex.list_feportsB = []
        for elt in portlist.findall("Port"):
            port = vplexPort.loadFromXML(elt)
            MyVPlex.list_feports.append(port)
            if port.PortName.find("-A0-FC") > 0:
                MyVPlex.list_feportsA.append(port)
            else:
                MyVPlex.list_feportsB.append(port)


        MyVPlex.list_storageA=[]
        for elt in XMLString.findall("Storage/ArrayList/Array"):
            MyVPlex.list_storageA.append(storageArray.loadFromXML(elt))
            #print(storageArray.loadFromXML(elt).toString())

        stvolumes = XMLString.find("StorageVolumes/thin-rebuild")
        MyVPlex.ClaimedCapacity = stvolumes.find("ClaimedCapacity").text.strip()
        MyVPlex.StorageVolumes = int(stvolumes.find("Count").text.strip())

        raid = XMLString.find("DeviceSummary/Raid0")
        MyVPlex.nbRaid0Devices = int(raid.find("NumberDevices").text)

        raid = XMLString.find("DeviceSummary/RaidC")
        MyVPlex.nbRaidCDevices = int(raid.find("NumberDevices").text)

        raid = XMLString.find("DeviceSummary/Raid1")
        MyVPlex.nbRaid1Devices = int(raid.find("NumberDevices").text)

        raid = XMLString.find("DeviceSummary/Distributed")
        MyVPlex.nbDistributedDevices = int(raid.find("ClusterNumberDistributedDevices").text)

        MyVPlex.nbRemoteFrom = int(raid.find("RemoteExportsFromThisCluster").text)
        MyVPlex.nbExportedVV = int(XMLString.find("NumberOfExportedVV").text)

        todo = XMLString.find("ConsistencyGroups/Local")
        MyVPlex.nbLocalCG = int(todo.find("NumberCGs").text)

        todo = XMLString.find("ConsistencyGroups/DistributedSync")
        MyVPlex.nbDistributedCG = int(todo.find("NumberCGs").text)

        chassis = XMLString.findall("ChassisList/Chassis")[0]
        MyVPlex.modelType = "UNK"
        if chassis.find("ChassisType").text.strip() == "VPL":
            MyVPlex.modelType="VS6"
        if chassis.find("ChassisType").text.strip() == "Argonaut":
            MyVPlex.modelType = "VS2"
        MyVPlex.seed_id = chassis.find("ChassisWWNSeed").text.strip()

        for iom in chassis.findall("IOModuleList/IOModule"):
            print(iom.find("Name").text.strip()+" - "+iom.find("Type").text.strip())

        return MyVPlex


def objectToXLS(Cell,chaine : str ,monObject : mesObjets):
    if str(cell.value).startswith(chaine):
        #
        # Manage Sym Data
        #
        Attributes = str(cell.value).replace(chaine, "")
        cell.value = monObject.getValue(Attributes)

def ListToXLS(feuille, Cell,chaine : str ,maListe : [mesObjets]):
    if str(cell.value).startswith(chaine):
        #
        # Manage Liste of disks
        #
        Attributes = str(cell.value).replace(chaine, "")
        row_x = cell.row
        column_y = cell.column
        for elt in maListe:
            #
            # on écrit de bas en haut.
            #
            celltoupd = feuille.cell(row=row_x, column=column_y)
            celltoupd.value = elt.getValue(Attributes)
            row_x = row_x + 1



logger.info("Start VPLEX")
FileToProcess=""

"""
Process the arguments
"""
parser = argparse.ArgumentParser(description="VplexXMLToExcel helps you to translate the configuration of a given VPLEX (local or metro) to an XLS File.\n\nThe program needs a openpyxl in your python installation.")
parser.add_argument('-file_name',help='Allow you to precise the file name (full path or not) of the XML file to parse')
args = parser.parse_args()

"""
Process the symapi_db parameter
"""
if args.file_name is not None:
    print("File Name : "+args.file_name)
    FileToProcess=args.file_name
else:
    logger.error("Missing pararemeters File_name leaving VplexXMLToExcel")
    print("Missing pararemeters File_name leaving VplexXMLToExcel")
    exit(1)

XMLStr=open(FileToProcess, 'r').read()
# clean \n
XMLStr = XMLStr.replace('\n', '')

#
# Make it XML
tableauXML = ET.fromstring(XMLStr)
versionXML = tableauXML.find("Version")
SystemID = tableauXML.find("SystemID").text.strip()


#
# Is this a Metro ?
TypeVPLEX = tableauXML.find("productType").text.strip()


print(TypeVPLEX)
list_vplex=[]

for nodeXML in tableauXML.findall("ClusterList/Cluster"):
    MyVPLEX = vplex.loadFromXML(nodeXML)
    if MyVPLEX.clusterTLA == SystemID:
        MyVPLEX.version = versionXML.find("ProductVersion").text.strip()
        MyVPLEX.siteID = tableauXML.find("CSISiteID").text.strip()
    list_vplex.append(MyVPLEX)
    print(MyVPLEX.toString())

#
# Copy XLS
#
dest_file_name=""
if TypeVPLEX == "Local":
    dest_file_name = list_vplex[0].clusterTLA + '.xlsx'
if TypeVPLEX == "Metro":
    dest_file_name=list_vplex[0].clusterTLA+"_"+list_vplex[1].clusterTLA+ '.xlsx'

copyfile("referenceVplexMetro.xlsx", dest_file_name)
#
# open the file and start to work
#
print("Populating data in  : " + dest_file_name)
tip = time.time()
classeur = openpyxl.load_workbook(dest_file_name)
for feuille_name in classeur.sheetnames:
    #
    # On parcourt les pages
    #
    feuille = classeur[feuille_name]
    for ligne in feuille.iter_rows():
        for cell in ligne:
            #
            # Analyse et travaille ici
            #
            objectToXLS(cell,"%%vplex1.",list_vplex[0])
            ListToXLS(feuille,cell,"%%listSto.vplex1.",list_vplex[0].list_storageA)
            ListToXLS(feuille, cell, "%%listFeDirA.vplex1.", list_vplex[0].list_feportsA)
            ListToXLS(feuille, cell, "%%listFeDirB.vplex1.", list_vplex[0].list_feportsB)
            if TypeVPLEX == "Metro":
                objectToXLS(cell, "%%vplex2.", list_vplex[1])
                ListToXLS(feuille, cell, "%%listSto.vplex2.", list_vplex[1].list_storageA)
                ListToXLS(feuille, cell, "%%listFeDirA.vplex2.", list_vplex[1].list_feportsA)
                ListToXLS(feuille, cell, "%%listFeDirB.vplex2.", list_vplex[1].list_feportsB)

    #
    # Save File
    #
    classeur.save(dest_file_name)
    top = time.time()
    print("Done in " + str(math.ceil(top - tip)) + " sec \n\n")


logger.info("End VPLEX")


