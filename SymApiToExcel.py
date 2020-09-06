"""
 SymApiToExcel

 Transforme la symapi d'une baie en fichier XML

 Author : Olivier Guyot

 Developped during off hours (vacations)
 Code is licences under GNU GPL v3

 TODO :
  Rajouter l'information sur le TDEV de montage d'un snap (ID destiantion). PAS D'idée pour le moment
  Rajouter les flgas sur les SG

"""

import argparse
import logging.config
import math
import os
import subprocess
import time
import xml.etree.ElementTree as ET
from shutil import copyfile

import openpyxl

"""
Initialize logging


File SymApiToExcel.logging mus be present and describe all availables logging configuration
Per default in the file :
2 loggers : 1 for console; one in file. (no rotation).
File is more for debug purposes.
In case of issue, truncate the log file or upgrade the logger level to Fatal

"""
logging.config.fileConfig('SymApiToExcel.logging')
logger = logging.getLogger('root')

"""
Initialize constants for the programs.
Mainly ou will find here SymCli commands

-out xml specified output in XML that will be parsed
variable %% %% is peace of text that will be later replace byt the parameter value. (mainly sid's)

"""
SymcfgList = 'symcfg list -v -output xml'
SymCfgListSrp = 'symcfg list -sid %%sid%% -srp -detail -out xml'
SymcfgEfficiency = 'symcfg -sid %%sid%% -srp -efficiency list -output xml'
SymcfgDemand = 'symcfg -sid %%sid%% list  -demand -v -tb -out xml'
SymCfgListTdev = 'symcfg -sid %%sid%% list -tdev -out xml'
SymDiskList = 'symdisk list -sid %%sid%% -out xml'
SymSGList = 'symsg list -v -sid %%sid%% -out xml'
SymDevShow = 'symdev show -sid %%sid%%  %%device%% -out xml '
SymDevList = 'symdev list -sid %%sid%%  -v -out xml '
SymCfgListMemory = 'symcfg -sid %%sid%%  list -memory -out xml'
SymCfgListFa = 'symcfg -sid %%sid%%  list -fa all -v -out xml'
SymCfgListRa = 'symcfg -sid %%sid%%  list -ra all -v -out xml'
SymSnapvxListe = 'symsnapvx -sid %%sid%% list -v -out xml'
SymSnapVXListDetails = 'symsnapvx -sid %%sid%% list -v -snapshot %%snap%% -dev 00001:FFFFF -detail -gb -out xml'
SymcfgListEmulation = 'symcfg -sid %%sid%% list -dir all -out xml'
SymcfgListPool = 'symcfg list -sid %%sid%% -pool -thin -out xml'
Supported_Platform = ['VMAX100K','VMAX200K','VMAX400K','VMAX250F', 'VMAX950F', 'VMAX450F', 'VMAX850F', 'PowerMax_8000', 'PowerMax_2000']




class mesObjets:
    """
    meObjets : mother class for all custom objects.

    you will find static methoods for code refactoring (why do it 10 times when you can write once and call many (find / findall)

    You will also find toString methods that transform an object to string listing all attributes to string.
    if an attribute is a list, then data are not printed.
    """


    def toString(self) -> str:
        """
        toString : put in a nice string all attributes of an object.

        :return: A string cotaining all attributes and values of an object
        """
        result = ""
        variables = self.__dict__.items()
        for variable, value in variables:
            if (str(variable).startswith("list_")):
                result = result + variable + " ====> LIST " + str(len(list(value))) + " Elements \n"
            else:
                result = result + variable + "  ====>  " + str(value) + "\n"
        return result

    def getValue(self, findV):
        """
        retrieve the value of a variable inside an object
        :param findV:
        :return: the value
        """
        variables = self.__dict__.items()
        for variable, value in variables:
            if variable == findV:
                return value
        return

    @staticmethod
    def _runfind(toRun) -> str:
        """
        Static method that execute an external command and retrieve its output into a String
        :param toRun: external command to run
        :return: a String containt the output of the command
        """
        logger.info("_runfind : " + toRun)
        liste = subprocess.check_output(toRun, shell=True)
        return liste



    @staticmethod
    def runFindall(toRun, toSearch) -> [ET.Element]:
        """
        Execute a command, convert output on XML and find all keys corresponding to the search
        :param toRun: output commant to Run
        :param toSearch: the XML treee to parse and retrieve
        :return: an XMLElement (kind of list).
        """
        tableauXML = ET.fromstring(mesObjets._runfind(toRun))
        return tableauXML.findall(toSearch)

    @staticmethod
    def runFind(toRun, toSearch) -> ET.Element:
        """
        Execute a command, convert output on XML and find 1 key corresponding to the search
        :param toRun: output commant to Run
        :param toSearch: the XML treee to parse and retrieve
        :return: an XMLElement (kind of list).
        """
        tableauXML = ET.fromstring(mesObjets._runfind(toRun))
        return tableauXML.find(toSearch)

    @staticmethod
    def ifNAtoInt(Value) -> int:
        """
        If value ="N/A", then return -1, if not return an int value of the value
        :param Value:
        :return:
        """
        result = 0
        if (Value == "N/A"):
            result = 0
        else:
            result = int(Value)

        return result

    @staticmethod
    def ifNAtoFloat(Value) -> float:
        """
        If value ="N/A", then return -1.0, if not return an float value of the value
        :param Value:
        :return:
        """
        result = 0.0
        if Value == "N/A":
            result = 0.0
        else:
            result = float(Value)

        return result


class emulation(mesObjets):
    """
    class for the FE ports (Frontend FC) FA Emulation

    Aim it to list all available Frontend Ports

    No special methods except :
    loadfromXML and loadfrom command (symdisk).
    loadfromXML do the mapping from XML to Object

    """
    id = ""
    type = ""
    symbolic = ""
    number = 0
    slot = 0
    status = ""
    cores = 0
    engine_num = 0
    ports = 0

    @staticmethod
    def loadSymmetrixFromXML(dirinfoXML):
        newEmulation = emulation()
        newEmulation.id = dirinfoXML.find("id").text
        newEmulation.type = dirinfoXML.find("type").text
        newEmulation.symbolic = dirinfoXML.find("symbolic").text
        newEmulation.number = int(dirinfoXML.find("number").text)
        newEmulation.slot = int(dirinfoXML.find("slot").text)
        newEmulation.status = dirinfoXML.find("status").text
        newEmulation.cores = int(dirinfoXML.find("cores").text)
        newEmulation.engine_num = int(dirinfoXML.find("engine_num").text)
        newEmulation.ports = int(dirinfoXML.find("ports").text)

        return newEmulation

    @staticmethod
    def loadFromCommand(sid) -> list:
        toRun = SymcfgListEmulation.replace('%%sid%%', sid)
        listEmulations = []
        for director in mesObjets.runFindall(toRun, 'Symmetrix/Director'):
            dir_info = director.find("Dir_Info")
            listEmulations.append(emulation.loadSymmetrixFromXML(dir_info))
        return listEmulations


class frontEndPorts(mesObjets):
    """
    class for the FE ports (Frontend FC) FA Emulation

    Aim it to list all available Frontend Ports

    No special methods except :
    loadfromXML and loadfrom command (symdisk).
    loadfromXML do the mapping from XML to Object

    """
    dir_name = ""
    port = 0
    port_wwn = ""
    port_status = ""
    negotiated_speed = 0
    maximum_speed = 0

    @staticmethod
    def loadSymmetrixFromXML(feXML,director_name):
        newFE = frontEndPorts()
        newFE.dir_name = director_name
        port_type=feXML.find("Port_Info")
        newFE.port = int(port_type.find("port").text)
        newFE.negotiated_speed = mesObjets.ifNAtoInt(port_type.find("negotiated_speed").text)
        newFE.maximum_speed = mesObjets.ifNAtoInt(port_type.find("maximum_speed").text)
        newFE.port_wwn = port_type.find("port_wwn").text
        newFE.port_status = port_type.find("port_status").text

        return newFE

    @staticmethod
    def loadFromCommand(sid) -> list:
        toRun = SymCfgListFa.replace('%%sid%%', sid)
        listFEPorts = []
        for director in mesObjets.runFindall(toRun, 'Symmetrix/Director'):
            dir_info = director.find("Dir_Info")
            dir_name = dir_info.find("symbolic").text
            for fe in director.findall("Port"):
                listFEPorts.append(frontEndPorts.loadSymmetrixFromXML(fe,dir_name))
        return listFEPorts

class replicationPorts(mesObjets):
    """
    class for the FE ports (Frontend FC) FA Emulation

    Aim it to list all available Frontend Ports

    No special methods except :
    loadfromXML and loadfrom command (symdisk).
    loadfromXML do the mapping from XML to Object

    """
    dir_name = ""
    port = 0
    port_status = ""
    negotiated_speed = 0
    maximum_speed = 0
    rdfList = ""

    @staticmethod
    def loadSymmetrixFromXML(feXML,director_name):
        newFE = replicationPorts()
        newFE.dir_name = director_name
        port_type=feXML.find("Port_Info")
        newFE.port = int(port_type.find("port").text)
        newFE.negotiated_speed = mesObjets.ifNAtoInt(port_type.find("negotiated_speed").text)
        newFE.maximum_speed = mesObjets.ifNAtoInt(port_type.find("maximum_speed").text)
        #newFE.port_wwn = port_type.find("port_wwn").text
        newFE.port_status = port_type.find("port_status").text
        newFE.rdfList=""
        for rdf in feXML.findall("RDF"):
            remote_symid = rdf.find("remote_symid").text
            ra_group_num = rdf.find("ra_group_num").text
            remote_ra_group_num = rdf.find("remote_ra_group_num").text
            newFE.rdfList=newFE.rdfList+"Remote : "+remote_symid+" (from RaGroup : "+ra_group_num+" to remote RaGroup : "+ remote_ra_group_num + ") / "
        return newFE

    @staticmethod
    def loadFromCommand(sid) -> list:
        toRun = SymCfgListRa.replace('%%sid%%', sid)
        listRAPorts = []
        for director in mesObjets.runFindall(toRun, 'Symmetrix/Director'):
            dir_info = director.find("Dir_Info")
            dir_name = dir_info.find("symbolic").text
            for fe in director.findall("Port"):
                listRAPorts.append(replicationPorts.loadSymmetrixFromXML(fe,dir_name))
        return listRAPorts


class disk(mesObjets):
    """
    class for the physical spindles

    Aim it to list all disks in the symmetrix, size, vendor and revision.

    No special methods except :
    loadfromXML and loadfrom command (symdisk).
    loadfromXML do the mapping from XML to Object

    """


    ident = ""
    da_number = ""
    disk_group = 0
    disk_group_name = ""
    disk_location = ""
    technology = ""
    vendor = ""
    revision = ""
    rated_gigabytes = ""

    @staticmethod
    def loadSymmetrixFromXML(Diskette):
        newDisk = disk()
        Disk = Diskette.find('Disk_Info')
        newDisk.ident = Disk.find("ident").text
        newDisk.da_number = Disk.find("da_number").text
        newDisk.disk_group = int(Disk.find("disk_group").text)
        newDisk.disk_group_name = Disk.find("disk_group_name").text
        newDisk.disk_location = Disk.find("disk_location").text
        newDisk.technology = Disk.find("technology").text
        newDisk.vendor = Disk.find("vendor").text
        newDisk.revision = Disk.find("revision").text
        newDisk.rated_gigabytes = Disk.find("rated_gigabytes").text

        return newDisk

    @staticmethod
    def loadFromCommand(sid) -> list:
        toRun = SymDiskList.replace('%%sid%%', sid)
        listDisks = []
        for galette in mesObjets.runFindall(toRun, 'Symmetrix/Disk'):
            listDisks.append(disk.loadSymmetrixFromXML(galette))
        return listDisks

class snapshotDetails(mesObjets):
    """
    class for the snaphosts header (general info)

    Aim it to list all available snapshots and volumes in it

    Additionnal infor for all generations will be done in anoterh class

    No special methods except :
    loadfromXML and loadfrom command (symdisk).
    loadfromXML do the mapping from XML to Object

    """
    source = ""
    snapshot_name = ""
    timestamp = ""
    generation = 0
    link = ""
    restore = ""
    failed  = ""
    error_reason = ""
    GCM = ""
    zDP = ""
    secured = ""
    expanded = ""
    total_snapshot_dev_size_gb = 0
    total_deltas_gb = 0
    non_shared_gb = 0
    expiration_date = 0

    @staticmethod
    def loadSymmetrixFromXML(snapshotXML):
        MySnap = snapshotDetails()
        MySnap.source = snapshotXML.find("source").text
        MySnap.snapshot_name = snapshotXML.find("snapshot_name").text
        MySnap.timestamp = snapshotXML.find("timestamp").text
        MySnap.generation = int(snapshotXML.find("generation").text)
        MySnap.link = snapshotXML.find("link").text
        MySnap.restore = snapshotXML.find("restore").text
        MySnap.failed = snapshotXML.find("failed").text
        MySnap.error_reason = snapshotXML.find("error_reason").text
        MySnap.GCM = snapshotXML.find("GCM").text
        MySnap.zDP = snapshotXML.find("zDP").text
        MySnap.secured = snapshotXML.find("secured").text
        MySnap.expanded = snapshotXML.find("expanded").text
        MySnap.total_snapshot_dev_size_gb = float(snapshotXML.find("total_snapshot_dev_size_gb").text)
        MySnap.total_deltas_gb = float(snapshotXML.find("total_deltas_gb").text)
        MySnap.non_shared_gb = float(snapshotXML.find("non_shared_gb").text)
        MySnap.expiration_date = snapshotXML.find("expiration_date").text

        return MySnap

    @staticmethod
    def loadFromCommand(sid,liste_snapH) -> list:
        listSnapshotDetails = []
        for snapH in liste_snapH:
            #
            # you loop on the snapshot headers (snap Name
            #
            toRun = SymSnapVXListDetails.replace('%%sid%%', sid)
            toRun = toRun.replace('%%snap%%', snapH.snapshot_name)
            for snapD in mesObjets.runFindall(toRun,"Symmetrix/Snapvx/Snapshot"):
                snapshotDetail = snapshotDetails.loadSymmetrixFromXML(snapD)
                listSnapshotDetails.append(snapshotDetail)
                snapH.total_snapshot_dev_size_gb = snapH.total_snapshot_dev_size_gb + snapshotDetail.total_snapshot_dev_size_gb
                snapH.total_deltas_gb = snapH.total_deltas_gb + snapshotDetail.total_deltas_gb
                snapH.non_shared_gb = snapH.non_shared_gb + snapshotDetail.non_shared_gb

        return listSnapshotDetails

class snapshotMaster(mesObjets):
    """
    class for the snaphosts header (general info)

    Aim it to list all available snapshots and volumes in it

    Additionnal infor for all generations will be done in anoterh class

    No special methods except :
    loadfromXML and loadfrom command (symdisk).
    loadfromXML do the mapping from XML to Object

    """
    source = ""
    snapshot_name = ""
    last_timestamp = ""
    num_generations = 0
    link = ""
    restore = ""
    failed = ""
    error_reason = ""
    GCM = ""
    zDP = ""
    secured = ""
    expanded = ""
    device_list = []
    total_snapshot_dev_size_gb = 0
    total_deltas_gb = 0
    non_shared_gb = 0


    @staticmethod
    def loadSymmetrixFromXML(snapshotXML):
        MySnap = snapshotMaster()
        MySnap.source = snapshotXML.find("source").text
        MySnap.snapshot_name = snapshotXML.find("snapshot_name").text
        MySnap.last_timestamp = snapshotXML.find("last_timestamp").text
        MySnap.num_generations = int(snapshotXML.find("num_generations").text)
        MySnap.link = snapshotXML.find("link").text
        MySnap.restore = snapshotXML.find("restore").text
        MySnap.failed = snapshotXML.find("failed").text
        MySnap.error_reason = snapshotXML.find("error_reason").text
        MySnap.GCM = snapshotXML.find("GCM").text
        MySnap.zDP = snapshotXML.find("zDP").text
        MySnap.secured = snapshotXML.find("secured").text
        MySnap.expanded = snapshotXML.find("expanded").text
        MySnap.device_list = []
        MySnap.device_list.append(MySnap.source)

        return MySnap

    @staticmethod
    def findSnapShotMaster(liste,name):
        for elt in liste:
            if elt.snapshot_name == name:
                return elt
        return None

    @staticmethod
    def loadFromCommand(sid) -> list:
        toRun = SymSnapvxListe.replace('%%sid%%', sid)
        listSnapshotMaster = []
        for snapshot in mesObjets.runFindall(toRun, 'Symmetrix/Snapvx/Snapshot'):
            snap = snapshotMaster.loadSymmetrixFromXML(snapshot)
            exist = snapshotMaster.findSnapShotMaster(listSnapshotMaster,snap.snapshot_name)
            if exist is None:
                listSnapshotMaster.append(snap)
            else:
                exist.source=exist.source+","+snap.source
                exist.device_list.append(snap.source)
        return listSnapshotMaster




class storageGroup(mesObjets):
    """
    class for the storageGroup (list of devices altogether
    A storage groups holds the compression flag (to be or not compressed

    This object has to be intialized after Tdevs, becase StorageGroups owns Tdevices (list).

    No special methods except :
    loadfromXML and loadfrom command (symsg).
    loadfromXML do the mapping from XML to Object

    """

    name = ""
    emulation = ""
    Masking_views = ""
    SLO_name = ""
    Compression = ""
    vp_saved_percent = ""
    compression_ratio = ""
    Num_of_GKS = 0
    HostIOLimit_status = ""
    HostIOLimit_max_mb_sec = ""
    HostIOLimit_max_io_sec = ""
    list_devices = []
    nbVolumes = 0
    size_presented_in_gb = 0
    size_allocated_in_gb = 0
    volumeList = ""

    @staticmethod
    def loadSymmetrixFromXML(sgsXML, paramlist_devices):
        sg = storageGroup()
        sginfo = sgsXML.find("SG_Info")
        sg.name = sginfo.find("name").text
        sg.emulation = sginfo.find("emulation").text
        sg.Masking_views = sginfo.find("Masking_views").text

        sg.SLO_name = sginfo.find("SLO_name").text
        sg.Compression = sginfo.find("Compression").text
        sg.vp_saved_percent = sginfo.find("vp_saved_percent").text
        sg.compression_ratio = sginfo.find("compression_ratio").text
        sg.Num_of_GKS = int(sginfo.find("Num_of_GKS").text)
        sg.HostIOLimit_status = sginfo.find("HostIOLimit_status").text
        sg.HostIOLimit_max_mb_sec = sginfo.find("HostIOLimit_max_mb_sec").text
        sg.HostIOLimit_max_io_sec = sginfo.find("HostIOLimit_max_io_sec").text

        dev_lists = sgsXML.find("DEVS_List")
        if dev_lists is None:
            #
            # Storage group has no volume
            #
            sg.nbVolumes = 0
            sg.size_presented_in_gb = 0
            sg.size_allocated_in_gb = 0
            sg.volumeList = ""
        else:
            #
            # Storage group has volumes
            #
            sg.nbVolumes = 0
            sg.size_presented_in_gb = 0
            sg.size_allocated_in_gb = 0
            sg.volumeList = ""
            #
            # Fetch all volumes from the device list
            #
            for device in dev_lists.findall("Device"):
                sg.nbVolumes = sg.nbVolumes + 1
                configuration = device.find("configuration").text
                volID = device.find("dev_name").text
                sg.volumeList = sg.volumeList + volID + ","
                for Tdev in paramlist_devices:
                    if (Tdev.dev_name == volID):
                        sg.list_devices.append(Tdev)
                        sg.size_presented_in_gb = sg.size_presented_in_gb + Tdev.total_tracks_gb
                        sg.size_allocated_in_gb = sg.size_allocated_in_gb + Tdev.alloc_tracks_gb
                        Tdev.configuration = configuration

        return sg

    @staticmethod
    def loadFromCommand(sid, paramlist_devices) -> list:
        toRun = SymSGList.replace('%%sid%%', sid)
        listSgs = []
        for sg in mesObjets.runFindall(toRun, 'SG'):
            listSgs.append(storageGroup.loadSymmetrixFromXML(sg, paramlist_devices))

        return listSgs



class tdev(mesObjets):
    """
        class for the Thin Devices (tdev)


        No special methods except :
        loadfromXML and loadfrom command (symdev and symcfg).
        loadfromXML do the mapping from XML to Object

        """
    dev_name = ""
    dev_emul = ""
    total_tracks_gb = 0
    alloc_tracks_gb = 0
    compression_ratio = ""
    tdev_status = ""
    configuration = ""
    emulation = ""
    encapsulated = ""
    encapsulated_wwn = ""
    encapsulated_array_id = ""
    encapsulated_device_name = ""
    status = ""
    snapvx_source = ""
    snapvx_target = ""
    wwn = ""
    ports = ""
    pair_state = ""
    suspend_state = ""
    consistency_state = ""
    paired_with_concurrent = ""
    paired_with_cascaded = ""
    remote_dev_name = ""
    remote_symid = ""
    remote_wwn = ""
    remote_state = ""
    rdf_mode = ""
    pair_state_1 = ""
    suspend_state_1 = ""
    consistency_state_1 = ""
    paired_with_concurrent_1 = ""
    paired_with_cascaded_1 = ""
    remote_dev_name_1 = ""
    remote_symid_1 = ""
    remote_wwn_1 = ""
    remote_state_1 = ""
    rdf_mode_1 = ""

    @staticmethod
    def findDetails(ID, listeDeviceDetails):
        for device in listeDeviceDetails:
            devinfo = device.find("Dev_Info")
            dev_name = devinfo.find("dev_name").text
            if dev_name == ID:
                return device
        logger.info("Device not found : " + ID)
        return ""

    @staticmethod
    def loadSymmetrixFromXML(device, listeDeviceDetails):
        newTdev = tdev()
        newTdev.dev_name = device.find("dev_name").text
        newTdev.dev_emul = device.find("dev_emul").text

        newTdev.total_tracks_gb = float(device.find("total_tracks_gb").text)
        newTdev.alloc_tracks_gb = float(device.find("alloc_tracks_gb").text)
        newTdev.compression_ratio = device.find("compression_ratio").text
        newTdev.tdev_status = device.find("tdev_status").text

        #
        # Work on addition info
        #
        details = tdev.findDetails(newTdev.dev_name, listeDeviceDetails)

        # Device not Found
        if details == "":
            logger.debug("Skip -- Device has no details")
            newTdev.status = "UNK"
            return newTdev

        Dev_Info = details.find("Dev_Info")

        newTdev.encapsulated = Dev_Info.find("encapsulated").text
        newTdev.encapsulated_wwn = Dev_Info.find("encapsulated_wwn").text
        newTdev.encapsulated_array_id = Dev_Info.find("encapsulated_array_id").text
        newTdev.encapsulated_device_name = Dev_Info.find("encapsulated_device_name").text
        newTdev.status = Dev_Info.find("status").text
        newTdev.snapvx_source = Dev_Info.find("snapvx_source").text
        newTdev.snapvx_target = Dev_Info.find("snapvx_target").text

        newTdev.emulation = Dev_Info.find("emulation").text


        Dev_Info = details.find("Device_External_Identity")
        newTdev.wwn = Dev_Info.find("wwn").text
        newTdev.ports = ""
        fe = Dev_Info.find("Front_End")
        if fe is not None:
            for port in fe.findall("Port"):
                newTdev.ports = newTdev.ports + port.find("director").text + "-" + port.find("port").text + ","

        rdfAll = details.findall("RDF")
        nbRDF = 0
        if rdfAll is not None:
            for rdf in rdfAll:
                rdf_info = rdf.find("RDF_Info")
                rdf_mode = rdf.find("Mode")
                remote = rdf.find("Remote")
                if nbRDF == 0:
                    newTdev.pair_state = rdf_info.find("pair_state").text
                    newTdev.suspend_state = rdf_info.find("suspend_state").text
                    newTdev.consistency_state = rdf_info.find("consistency_state").text
                    newTdev.paired_with_concurrent = rdf_info.find("paired_with_concurrent").text
                    newTdev.paired_with_cascaded = rdf_info.find("paired_with_cascaded").text
                    newTdev.rdf_mode = rdf_mode.find("mode").text
                    newTdev.remote_dev_name = remote.find("dev_name").text
                    newTdev.remote_symid = remote.find("remote_symid").text
                    newTdev.remote_wwn = remote.find("wwn").text
                    newTdev.remote_state = remote.find("state").text
                else:
                    newTdev.pair_state_1 = rdf_info.find("pair_state").text
                    newTdev.suspend_state_1 = rdf_info.find("suspend_state").text
                    newTdev.consistency_state_1 = rdf_info.find("consistency_state").text
                    newTdev.paired_with_concurrent_1 = rdf_info.find("paired_with_concurrent").text
                    newTdev.paired_with_cascaded_1 = rdf_info.find("paired_with_cascaded").text
                    newTdev.rdf_mode_1 = rdf_mode.find("mode").text
                    newTdev.remote_dev_name_1 = remote.find("dev_name").text
                    newTdev.remote_symid_1 = remote.find("remote_symid").text
                    newTdev.remote_wwn_1 = remote.find("wwn").text
                    newTdev.remote_state_1 = remote.find("state").text
                nbRDF = nbRDF+1
                if nbRDF > 2:
                    print(newTdev.dev_name+" has "+str(nbRDF)+" SRDF")

        return newTdev

    @staticmethod
    def loadFromCommand(sid) -> list:
        toRun = SymDevList.replace('%%sid%%', sid)
        listeDetailsDevicesXML = mesObjets.runFindall(toRun, 'Symmetrix/Device')

        logger.info("NB elt : " + str(len(listeDetailsDevicesXML)))

        toRun = SymCfgListTdev.replace('%%sid%%', sid)
        listDdevices = []
        listeDevicesXML = mesObjets.runFindall(toRun, 'Symmetrix/ThinDevs/Device')
        logger.info("NB elt : " + str(len(listeDevicesXML)))

        for device in listeDevicesXML:
            listDdevices.append(tdev.loadSymmetrixFromXML(device, listeDetailsDevicesXML))

        return listDdevices


#
# Internal classes
# to move later
#
class symmetrix(mesObjets):
    """
    Top class object
    The symmetrix is the top object. it has attributes (SID, ....) but also list of items (Snaps, ....)
    """
    symid = ""
    attachment = ""
    product_model = ""
    disks = 0
    hot_spares = 0
    patch_level = ""
    raid_level = ""
    srp_name = ""
    vp_efficiency = ""
    snapshot_efficiency = ""
    data_reduction = ""
    drr_enabled_pct = 0
    SRP_efficiency = ""
    effective_used_cap_percent = 0
    usable_capacity_tb = 0.0
    used_capacity_tb = 0.0
    free_capacity_tb = 0.0
    subscribed_capacity_tb = 0.0
    user_used_capacity_tb = 0.0
    system_used_capacity_tb = 0.0
    temp_used_capacity_tb = 0.0
    array_meta_data_used_percent = 0
    repl_meta_data_used_percent = 0
    fe_meta_data_used_percent = 0
    be_meta_data_used_percent = 0
    snapshot_capacity_tb = 0.0
    snapshot_cap_nonshared_tb = 0.0
    snapshot_cap_shared_tb = 0.0
    snapshot_cap_modified_percent = 0
    list_disks = []
    list_devices = []
    list_sgs = []
    list_fes = []
    list_ras = []
    nb_engine = 0
    nb_cache_raw_tb = 0.0
    list_sm = []
    list_sd = []
    list_emu = []

    @staticmethod
    def loadSymmetrixFromXML(symm):
        newSymmtrix = symmetrix()

        #
        # go thru SymmInfo for generic Info.
        #
        symmInfo = symm.find('Symm_Info')

        newSymmtrix.symid = symmInfo.find("symid").text
        newSymmtrix.attachment = symmInfo.find("attachment").text
        newSymmtrix.product_model = symmInfo.find("product_model").text
        newSymmtrix.disks = int(symmInfo.find("disks").text)
        newSymmtrix.hot_spares = int(symmInfo.find("hot_spares").text)

        #
        # go thru Enginuity for code level
        #
        symmEnginuity = symm.find('Enginuity')

        newSymmtrix.patch_level = symmEnginuity.find("patch_level").text

        if newSymmtrix.product_model not in Supported_Platform:
            logger.error("Plateform not supported : " + newSymmtrix.product_model)
            return newSymmtrix

        #
        # go thru Flags for Raid Level
        symmFlags = symm.find('Flags')

        if symmFlags.find("raid_5").text == "N/A":
            newSymmtrix.raid_level = symmFlags.find("raid_6").text
        else:
            newSymmtrix.raid_level = symmFlags.find("raid_5").text

        #
        # Get efficiency data
        #
        toRun = SymcfgEfficiency.replace('%%sid%%', newSymmtrix.symid)
        srp = mesObjets.runFind(toRun, "Symmetrix/SRP/SRP_Info")
        newSymmtrix.srp_name = srp.find("name").text

        vp_eff = srp.find('vp_efficiency')
        newSymmtrix.vp_efficiency = vp_eff.find("overall_ratio").text

        snap_eff = srp.find('snapshot_efficiency')
        newSymmtrix.snapshot_efficiency = snap_eff.find("overall_ratio").text

        drr_eff = srp.find('data_reduction')
        newSymmtrix.data_reduction = drr_eff.find("ratio").text
        newSymmtrix.drr_enabled_pct = mesObjets.ifNAtoInt(drr_eff.find("enabled_percent").text)

        ovr_eff = srp.find('SRP_efficiency')
        newSymmtrix.SRP_efficiency = ovr_eff.find("overall_ratio").text

        #
        # Get The demand info
        #
        toRun = SymcfgDemand.replace('%%sid%%', newSymmtrix.symid)
        srp = mesObjets.runFind(toRun, "Symmetrix")

        if newSymmtrix.product_model in ['VMAX100K','VMAX200K','VMAX400K']:
            """
            If VMAX3, data are not gathered at the same pliace
            Some values are not also present, so put 0 in order to avaid later issues.
            """
            toRun = SymCfgListSrp.replace('%%sid%%', newSymmtrix.symid)
            srp = mesObjets.runFind(toRun, "Symmetrix/SRP/SRP_Info")

            newSymmtrix.usable_capacity_tb = mesObjets.ifNAtoFloat(srp.find("usable_capacity_terabytes").text)
            newSymmtrix.user_used_capacity_tb = 0
            newSymmtrix.free_capacity_tb = mesObjets.ifNAtoFloat(srp.find("free_capacity_terabytes").text)
            newSymmtrix.subscribed_capacity_tb = mesObjets.ifNAtoFloat(srp.find("subscribed_capacity_terabytes").text)
            newSymmtrix.system_used_capacity_tb = 0
            newSymmtrix.used_capacity_tb = newSymmtrix.usable_capacity_tb - newSymmtrix.free_capacity_tb
            newSymmtrix.temp_used_capacity_tb = 0
            newSymmtrix.array_meta_data_used_percent = 0
            newSymmtrix.repl_meta_data_used_percent = 0
            newSymmtrix.fe_meta_data_used_percent = 0
            newSymmtrix.be_meta_data_used_percent = 0
            newSymmtrix.snapshot_capacity_tb = 0
            newSymmtrix.snapshot_cap_nonshared_tb = 0
            newSymmtrix.snapshot_cap_shared_tb = 0
            newSymmtrix.snapshot_cap_modified_percent = 0
            newSymmtrix.effective_used_cap_percent = (newSymmtrix.used_capacity_tb * 100)//newSymmtrix.usable_capacity_tb
        else:
            #
            # This is a VMAX AFA or a PowerMax, so it has the requirent info
            #
            newSymmtrix.effective_used_cap_percent = mesObjets.ifNAtoInt(srp.find("effective_used_cap_percent").text)
            newSymmtrix.usable_capacity_tb = mesObjets.ifNAtoFloat(srp.find("usable_capacity_tb").text)
            newSymmtrix.user_used_capacity_tb = mesObjets.ifNAtoFloat(srp.find("user_used_capacity_tb").text)
            #newSymmtrix.used_capacity_tb = mesObjets.ifNAtoFloat(srp.find("used_capacity_tb").text)
            # Used capcity in the output XML is messy
            # change it to go to list pool
            newSymmtrix.subscribed_capacity_tb = mesObjets.ifNAtoFloat(srp.find("subscribed_capacity_tb").text)
            newSymmtrix.system_used_capacity_tb = mesObjets.ifNAtoFloat(srp.find("system_used_capacity_tb").text)
            newSymmtrix.temp_used_capacity_tb = mesObjets.ifNAtoFloat(srp.find("temp_used_capacity_tb").text)
            newSymmtrix.array_meta_data_used_percent = mesObjets.ifNAtoInt(srp.find("array_meta_data_used_percent").text)
            newSymmtrix.repl_meta_data_used_percent = mesObjets.ifNAtoInt(srp.find("repl_meta_data_used_percent").text)
            newSymmtrix.fe_meta_data_used_percent = mesObjets.ifNAtoInt(srp.find("fe_meta_data_used_percent").text)
            newSymmtrix.be_meta_data_used_percent = mesObjets.ifNAtoInt(srp.find("be_meta_data_used_percent").text)
            newSymmtrix.snapshot_capacity_tb = mesObjets.ifNAtoFloat(srp.find("snapshot_capacity_tb").text)
            newSymmtrix.snapshot_cap_nonshared_tb = mesObjets.ifNAtoFloat(srp.find("snapshot_cap_nonshared_tb").text)
            newSymmtrix.snapshot_cap_shared_tb = mesObjets.ifNAtoFloat(srp.find("snapshot_cap_shared_tb").text)
            newSymmtrix.snapshot_cap_modified_percent = mesObjets.ifNAtoInt(srp.find("snapshot_cap_modified_percent").text)
            #
            #
            #
            #
            toRun = SymcfgListPool.replace('%%sid%%',newSymmtrix.symid)
            listpool = mesObjets.runFind(toRun,"Symmetrix/Totals")
            newSymmtrix.used_capacity_tb = listpool.find("total_used_tracks_tb").text
            newSymmtrix.free_capacity_tb = listpool.find("total_free_tracks_tb").text
            newSymmtrix.user_used_capacity_tb = float(newSymmtrix.used_capacity_tb) - newSymmtrix.system_used_capacity_tb - newSymmtrix.temp_used_capacity_tb - newSymmtrix.snapshot_cap_nonshared_tb




        #
        # Load Disks
        #
        newSymmtrix.list_disks = disk.loadFromCommand(newSymmtrix.symid)

        #
        # get Engines and cache
        #
        toRun = SymCfgListMemory.replace('%%sid%%', newSymmtrix.symid)
        memory = mesObjets.runFind(toRun, "Symmetrix/Symm_Info")

        newSymmtrix.nb_engine = int(memory.find("total_mem_boards").text) // 2

        memory = mesObjets.runFindall(toRun, "Symmetrix/Memory_Board")
        newSymmtrix.nb_cache_raw_tb=0.0
        for board in memory:
            #
            # Welcome in try and guess ...
            #
            val = float(board.find("capacity_in_mb").text)
            if val > 800000:
                # 1TB cache director
                newSymmtrix.nb_cache_raw_tb=newSymmtrix.nb_cache_raw_tb+1
            elif val > 400000:
                # 512GB cache director
                newSymmtrix.nb_cache_raw_tb = newSymmtrix.nb_cache_raw_tb + 0.5
            else:
                # 256 GB cache director
                newSymmtrix.nb_cache_raw_tb = newSymmtrix.nb_cache_raw_tb + 0.25

        #
        # Load devices
        #
        newSymmtrix.list_devices = tdev.loadFromCommand(newSymmtrix.symid)

        #
        # Load SG's
        #
        newSymmtrix.list_sgs = storageGroup.loadFromCommand(newSymmtrix.symid, newSymmtrix.list_devices)


        #
        # grab the FA ports (frontend FC emulation)
        #
        newSymmtrix.list_fes=frontEndPorts.loadFromCommand(newSymmtrix.symid)

        #
        # grab the RE / RF pors (frontend FC emulation)
        #
        newSymmtrix.list_ras = replicationPorts.loadFromCommand(newSymmtrix.symid)

        #
        # grab the snapshot masters
        #

        newSymmtrix.list_sm = snapshotMaster.loadFromCommand(newSymmtrix.symid)
        newSymmtrix.list_sd = snapshotDetails.loadFromCommand(newSymmtrix.symid,newSymmtrix.list_sm)

        #
        # liste des emulations
        #
        newSymmtrix.list_emu = emulation.loadFromCommand(newSymmtrix.symid)


        return newSymmtrix

class smallSym(mesObjets):
    symmID = ""
    attachement = ""
    product_model = ""

    @staticmethod
    def loadFromCommand() -> list:
        symList = []
        for symm in mesObjets.runFindall(SymcfgList, 'Symmetrix'):

            mySym = smallSym()
            symminfo = symm.find("Symm_Info")
            mySym.symmID = symminfo.find("symid").text
            mySym.attachement = symminfo.find("attachment").text
            mySym.product_model = symminfo.find("product_model").text
            if mySym.product_model in Supported_Platform:
                symList.append(mySym)
            else:
                #print("skip unsupported system "+mySym.symmID+" / "+ mySym.product_model)
                logger.info("skip unsupported system "+mySym.symmID+" / "+ mySym.product_model)
        return symList




#
# Main
#


def objectToXLS(Cell,chaine : str ,monObject : mesObjets):
    if str(cell.value).startswith(chaine):
        #
        # Manage Sym Data
        #
        Attributes = str(cell.value).replace(chaine, "")
        cell.value = monObject.getValue(Attributes)

def ListToXLS(feuille, Cell,chaine : str ,maListe : [mesObjets]):
    if str(cell.value).startswith(chaine):
        #
        # Manage Liste of disks
        #
        Attributes = str(cell.value).replace(chaine, "")
        row_x = cell.row
        column_y = cell.column
        for elt in maListe:
            #
            # on écrit de bas en haut.
            #
            celltoupd = feuille.cell(row=row_x, column=column_y)
            celltoupd.value = elt.getValue(Attributes)
            row_x = row_x + 1

def whichSID() -> list:
    """
    Determines which SID shoulbe treated by this programm
    It lists all available SID's in the symapi
    List them on screen
    and manage to get the ID you want to use
    ALL = All ID's
    QUIT = exit(0) of the program

    :return: a list of SymmID
    """
    list_sym=[]
    """
    Let's start
    Step 1 : Gather all available (and supported) SID's
    Step 2 : Print Them
    Step 3 : loop until you get a decent answer

    """
    listSymm = smallSym.loadFromCommand()
    print("List of discovered systems : ")

    pos = 0
    answer_list = []
    answer_list.append("ALL")
    answer_list.append("QUIT")
    while pos < len(listSymm):
        answer_list.append((str(pos)))
        pos = pos + 1

    pos = 0
    for smallsym in listSymm:
        print(str(pos) + ' - ' + smallsym.symmID + ' (' + smallsym.attachement + ') - ' + smallsym.product_model)
        pos = pos + 1

    print("")
    print("please enter the system to process (0 to " + str(len(listSymm) - 1) + ") or ALL or QUIT")
    answer = input("which system to process : ")
    answer = answer.upper()
    while answer not in answer_list:
        answer = input("which system to process : ")
        answer = answer.upper()

    """
    Now we have got an answer, let's process the answer
    """
    if answer == 'QUIT':
        print("")
        logger.info("This is the end")
        exit(0)

    list_sym = []

    if answer == "ALL":
        for smallsym in listSymm:
            list_sym.append(smallsym.symmID)
    else:
        list_sym.append(listSymm[int(answer)].symmID)
    return list_sym

"""

MAIN STARTS HERE

"""
logger.info("Start")
"""
Process the arguments
"""
parser = argparse.ArgumentParser(description="SympApiToExcel helps you to translate the configuration of a given Symmetrix to an XLS File.\n\nThe program needs to have symcli 9.1 installed and in your path, as well as a openpywl in your python installation.")
parser.add_argument('-sid',help='Allow you to precise a SID (needs to be fully precise as in the symapi)')
parser.add_argument('-symapi_dir',help="allow you to precise a directory whe the symapi_db's are located, symapi_dbs should be in the form symapi*.bin",type=str)
parser.add_argument('-symapi_db',help="allow you to precise a precise SYMAPI_DB.bin",type=str)
parser.add_argument('-all',help="will run against all SYMIDs in the symapi_db",action="store_true")
parser.add_argument('-local',help="will run against all local SYMIDs in the symapi_db",action="store_true")
args = parser.parse_args()


"""
Update system Variables for offline symcli.
"""
my_vars = {'SYMCLI_OFFLINE': "1"}
os.environ.update(my_vars)
my_vars = {'SYMCLI_SNAPVX_LIST_OFFLINE': "enabled"}
os.environ.update(my_vars)

"""
Process the symapi_db parameter
"""
if args.symapi_db is not None:
    print("Parameter : "+args.symapi_db)
    my_vars = {'SYMCLI_DB_FILE': args.symapi_db}
    os.environ.update(my_vars)



"""
Process the symapi_dir
"""
if args.symapi_dir is not None:
    print("Parameter : "+args.symapi_dir)
    liste_symapi=[]
    for file in [f for f in os.listdir(args.symapi_dir)]:
        if file.lower().startswith("symapi"):
            if file.lower().endswith(".bin"):
                liste_symapi.append(file)
    pos = 0
    answer_list = []
    answer_list.append("QUIT")
    while pos < len(liste_symapi):
        answer_list.append((str(pos)))
        pos = pos + 1
    pos = 0
    for file in liste_symapi:
        print(str(pos)+" - "+file)
        pos = pos + 1

    print("")
    print("please enter the id of the symapi to process (0 to " + str(len(liste_symapi) - 1) + ") or QUIT")
    answer = input("which symapi id : ")
    answer = answer.upper()
    while answer not in answer_list:
        answer = input("which symapi id : ")
        answer=answer.upper()

    """
    Now we have got an answer, let's process the answer
    """
    if answer == 'QUIT':
        print("")
        logger.info("This is the end")
        exit(0)
    print("")
    Symapifile=args.symapi_dir+os.path.sep+liste_symapi[int(answer)]
    print("Selected Symapi is : " + Symapifile)
    my_vars = {'SYMCLI_DB_FILE': Symapifile}
    os.environ.update(my_vars)
    print("")


list_sym = []
if args.all:
    print("Flag All selected.")
    print("The following symmetrix will be audited :")
    listSymm = smallSym.loadFromCommand()
    for symm in listSymm:
        print(symm.symmID)
        list_sym.append(symm.symmID)

if args.local:
    print("Flag local selected.")
    print("The following symmetrix will be audited :")
    listSymm = smallSym.loadFromCommand()
    for symm in listSymm:
        if symm.attachement == "Local":
            print(symm.symmID)
            list_sym.append(symm.symmID)


"""
Manage sid value
"""

if args.sid is not None:
    print("Parameter : "+args.sid)
    list_sym.append(args.sid)
else:
    if len(list_sym) == 0:
        list_sym = whichSID()

print("")
print("")

"""
Restart looping on all symmetrix and check if it is in the process list. If not skipp and try next one
"""
for symm in mesObjets.runFindall(SymcfgList, 'Symmetrix'):
    symminfo = symm.find("Symm_Info")
    curr_symmID = symminfo.find("symid").text

    if curr_symmID not in list_sym:
        """
        Not in the list to process
        """
        continue

    product_model = symminfo.find("product_model").text

    if product_model not in Supported_Platform:
        """
        Not Supported.
        """
        print(product_model+" is not supported")
        continue

    print("loading symmetrix data from Symapi  : "+curr_symmID)
    tip = time.time()
    MySymm = symmetrix.loadSymmetrixFromXML(symm)
    top = time.time()
    print("Done in "+str(math.ceil(top-tip))+" sec")
    logger.info(MySymm.toString())

    """
    Let's manage the output.
    """
    #
    # Copy XLS
    #
    copyfile("reference.xlsx", MySymm.symid + '.xlsx')
    #
    # open the file and start to work
    #
    print("Populating data in  : " + MySymm.symid + '.xlsx')
    tip = time.time()
    classeur = openpyxl.load_workbook(MySymm.symid + '.xlsx')
    for feuille_name in classeur.sheetnames:
        #
        # On parcourt les pages
        #
        feuille = classeur[feuille_name]
        for ligne in feuille.iter_rows():
            for cell in ligne:
                #
                # Analyse et travaille ici
                #
                objectToXLS(cell,"%%sym.",MySymm)
                ListToXLS(feuille,cell,"%%list.disks.",MySymm.list_disks)
                ListToXLS(feuille, cell, "%%list.tdevs.", MySymm.list_devices)
                ListToXLS(feuille, cell, "%%list.sgs.", MySymm.list_sgs)
                ListToXLS(feuille, cell, "%%list.fes.", MySymm.list_fes)
                ListToXLS(feuille, cell, "%%list.ras.", MySymm.list_ras)
                ListToXLS(feuille, cell, "%%list.sm.", MySymm.list_sm)
                ListToXLS(feuille, cell, "%%list.sd.", MySymm.list_sd)
                ListToXLS(feuille, cell, "%%list.emu.", MySymm.list_emu)

    #
    # Save File
    #
    classeur.save(MySymm.symid + '.xlsx')
    top = time.time()
    print("Done in " + str(math.ceil(top - tip)) + " sec \n\n")

    #    for child in symm:
    #        print(child.tag, child.attrib)


print("Finally finished")
logger.info("End")
